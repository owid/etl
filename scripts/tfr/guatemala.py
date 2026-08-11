"""Guatemala: registered births by age of mother over INE's population projections.

INE publishes no annual fertility rate from the registry — its bulletin headlines a crude birth rate
and rates for teenagers — but it does publish the raw birth records, one row per birth, with the
mother's age and the year the birth happened. So the rate is built from scratch here.

Two of INE's own figures corroborate the result: its 2018 census estimate of 2.7 for 2018-19 against
our 2.6, and about 2.2 for 2022 from its maternal and child health survey against our 2.23.

Births are dated to the year they occurred, with a six-month window for late registration. That window
is not always enough. The 2024 file holds 297,408 births to women 15-49, 12.7% fewer than 2023, after
four years that moved by about a percent either way — and reporting on the 2025 figures shows part of
that drop reversing. A one-year discontinuity that large, followed by a rebound, is what an
under-registered year looks like rather than a fall in fertility, so the series stops at the last year
whose file looks settled. Advance LAST_COMPLETE when a later file stops growing.
"""

import os
import re
import subprocess

import pandas as pd

DATA = os.path.join(os.path.dirname(__file__), "data", "gt")
# the newest year whose birth file looks settled rather than still filling in; see the note above
LAST_COMPLETE = 2023
PORTAL = "https://datos.ine.gob.gt/dataset/96888fc0-5ced-4a58-bfd0-2a6a292e3208/resource"
BIRTHS = {
    2018: f"{PORTAL}/41b95ebc-35c6-43d3-86db-4b85998cda04/download/nacimientos-2018.xlsx",
    2019: f"{PORTAL}/c32be289-eceb-4264-a570-1cecef37635c/download/nacimientos-2019.xlsx",
    2020: f"{PORTAL}/146d86fb-3736-4536-a3fa-09a047b7e601/download/nacimientos-2020.xlsx",
    2021: f"{PORTAL}/25bfa38b-0d32-4b4f-aed9-6ae57eae38ed/download/nacimientos-2021.xlsx",
    2022: f"{PORTAL}/a7d24194-227e-4cbe-9fd8-d7ddb92eb2b3/download/nacimientos-2022.xlsx",
    2023: f"{PORTAL}/ecfb6233-d695-46ee-be95-386d25aebb14/download/bdnacimientos2023.xlsx",
    2024: f"{PORTAL}/14f54638-cf4a-4476-9a45-e03ab2baa3c3/download/nacimientos2024_da.xlsx",
}
WOMEN = ("https://datos.ine.gob.gt/dataset/23bbe8f8-ec9b-49c9-bcca-de78bec4efc0/resource/"
         "56b97bc0-5624-408d-a136-3931dd05610a/download/estimaciones-y-proyecciones-de-la-"
         "poblacion-total-por-municipio-segun-sexo-y-edad.-2015-2030.xlsx")
BANDS = [(15, 19), (20, 24), (25, 29), (30, 34), (35, 39), (40, 44), (45, 49)]
UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
FEMALE = 2          # the projections file codes sex 0 total, 1 male, 2 female
UNKNOWN_AGE = 999   # the births file codes an unrecorded age as 999


def _get(url, name):
    """Download once. These are tens of megabytes each, so they are kept."""
    path = os.path.join(DATA, name)
    if not os.path.exists(path):
        os.makedirs(DATA, exist_ok=True)
        subprocess.run(["curl", "-sL", "--fail", "-A", UA, "-m", "900", "-o", path, url], check=True)
    return path


def _band(age):
    for lo, hi in BANDS:
        if lo <= age <= hi:
            return (lo, hi)
    return None


def _births_file(path):
    """{band: births} for the year, counted from the individual records.

    The file is one row per birth and too large to hold in memory, so it is read a row at a time.
    Column names vary a little between editions, so they are matched rather than assumed.
    """
    from openpyxl import load_workbook

    book = load_workbook(path, read_only=True, data_only=True)
    sheet = book[book.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = [str(h).strip().lower() if h is not None else "" for h in next(rows)]
    age_col = next(i for i, h in enumerate(header) if h.startswith("edadm"))
    out = {}
    for row in rows:
        raw = row[age_col]
        age = pd.to_numeric(raw, errors="coerce")
        if pd.isna(age) or int(age) == UNKNOWN_AGE:
            continue
        band = _band(int(age))
        if band:
            out[band] = out.get(band, 0.0) + 1.0
    book.close()
    return out


def _births():
    return {year: _births_file(_get(url, f"births_{year}.xlsx")) for year, url in BIRTHS.items()}


def _women():
    """{year: {band: women}} from the municipal projections, summed to the whole country.

    Ages run down the rows as single years with an open top group, and years across the columns.
    """
    d = pd.read_excel(_get(WOMEN, "women.xlsx"), sheet_name=0)
    d.columns = [str(c).strip() for c in d.columns]
    sex = next(c for c in d.columns if c.upper().startswith("SEXO"))
    age = next(c for c in d.columns if c.upper().startswith("EDAD"))
    years = {c: int(c) for c in d.columns if re.fullmatch(r"20\d{2}", str(c))}
    d = d[pd.to_numeric(d[sex], errors="coerce") == FEMALE].copy()
    d["band"] = [_band(int(a)) if pd.notna(pd.to_numeric(a, errors="coerce")) else None
                 for a in pd.to_numeric(d[age], errors="coerce")]
    d = d[d.band.notna()]
    out = {}
    for col, year in years.items():
        by_band = d.groupby("band", observed=True)[col].sum()
        out[year] = {band: float(v) for band, v in by_band.items()}
    return out


def guatemala_tfr():
    births, women = _births(), _women()
    rows = []
    for year in sorted(set(births) & set(women)):
        if year > LAST_COMPLETE:
            continue
        b, w = births[year], women[year]
        if all(b.get(x) and w.get(x) for x in BANDS):
            rows.append({"year": year, "value": sum(b[x] / w[x] for x in BANDS) * 5})
    return pd.DataFrame(rows)


def guatemala_detail(year):
    births, women = _births().get(year), _women().get(year)
    if not births or not women:
        return None
    return {b: {"births": births[b], "women": women[b]} for b in BANDS if b in births and b in women}


if __name__ == "__main__":
    print(guatemala_tfr().to_string(index=False))
    print("INE's projection assumes 2.44 for 2022-23 and 2.33 for 2024-25")
